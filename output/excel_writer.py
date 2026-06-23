"""Styled Excel output with flagged cells for non-OK fields."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config.field_config import FieldConfig
from models import DocumentResult, FieldResult, FieldStatus
from output.formatting import (
    display_field_value,
    is_document_failed,
    overall_status_display,
)

_FILL_FAILED = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
_FILL_LOW_CONFIDENCE = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
_FILL_NOT_FOUND = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
_FILL_DOCUMENT_FAILED = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
_HEADER_FONT = Font(bold=True)

_STATUS_FILLS = {
    FieldStatus.FAILED_VALIDATION: _FILL_FAILED,
    FieldStatus.LOW_CONFIDENCE: _FILL_LOW_CONFIDENCE,
    FieldStatus.NOT_FOUND: _FILL_NOT_FOUND,
}


def _header_row(field_configs: list[FieldConfig]) -> list[str]:
    return (
        ["Source File"]
        + [field.display_label for field in field_configs]
        + ["Overall Status"]
    )


def _column_widths(headers: list[str]) -> list[int]:
    widths: list[int] = []
    for header in headers:
        if header == "Source File":
            widths.append(28)
        elif header == "Overall Status":
            widths.append(24)
        else:
            widths.append(max(14, min(len(header) + 4, 40)))
    return widths


def _apply_field_cell(
    cell,
    field_result: FieldResult,
    *,
    document_failed: bool,
) -> None:
    if document_failed:
        cell.value = None
        cell.fill = _FILL_DOCUMENT_FAILED
        return

    display_value = display_field_value(field_result)
    cell.value = display_value

    fill = _STATUS_FILLS.get(field_result.status)
    if fill is not None:
        cell.fill = fill

    if field_result.status != FieldStatus.OK and field_result.reason:
        cell.comment = Comment(field_result.reason, "validator")


def write_excel(
    results: list[DocumentResult],
    field_configs: list[FieldConfig],
    output_path: Path | str,
) -> Path:
    """Write one row per document to a styled .xlsx file."""
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invoices"

    headers = _header_row(field_configs)
    worksheet.append(headers)

    for column_index, width in enumerate(_column_widths(headers), start=1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = width

    for cell in worksheet[1]:
        cell.font = _HEADER_FONT

    for document in results:
        row: list[object] = [document.source_filename]
        document_failed = is_document_failed(document)

        for field_config in field_configs:
            field_result = document.fields.get(field_config.key)
            if field_result is None:
                field_result = FieldResult(status=FieldStatus.NOT_FOUND)
            row.append(display_field_value(field_result) if not document_failed else None)

        row.append(overall_status_display(document))
        worksheet.append(row)

        row_index = worksheet.max_row
        for field_offset, field_config in enumerate(field_configs, start=2):
            field_result = document.fields.get(field_config.key)
            if field_result is None:
                field_result = FieldResult(status=FieldStatus.NOT_FOUND)
            cell = worksheet.cell(row=row_index, column=field_offset)
            _apply_field_cell(cell, field_result, document_failed=document_failed)

        if document_failed:
            overall_cell = worksheet.cell(row=row_index, column=len(headers))
            if document.error_message:
                overall_cell.comment = Comment(document.error_message, "pipeline")

    worksheet.freeze_panes = "A2"
    workbook.save(out_path)
    return out_path
