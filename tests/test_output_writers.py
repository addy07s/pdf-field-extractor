"""Tests for Excel and CSV output writers."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from config.field_config import load_field_configs
from models import DocumentResult, DocumentStatus, FieldResult, FieldStatus
from output.csv_writer import write_csv
from output.excel_writer import write_excel

SMOKE_FIELDS = {
    "company_name": FieldResult(value="ADITYA MULTIMEDIA AND ENTERTAINMENT", status=FieldStatus.OK),
    "invoice_number": FieldResult(value="AME/2", status=FieldStatus.OK),
    "invoice_date": FieldResult(
        value={"original": "23/04/2026", "normalized": "2026-04-23"},
        status=FieldStatus.OK,
    ),
    "supplier_gstin": FieldResult(value="24ABMFA4190N1Z2", status=FieldStatus.OK),
    "recipient_gstin": FieldResult(status=FieldStatus.NOT_FOUND, reason="field not extracted"),
    "pan": FieldResult(value="ABMFA4190N", status=FieldStatus.OK),
    "description": FieldResult(
        value="ROYALTY INCOME",
        status=FieldStatus.LOW_CONFIDENCE,
        reason="value not found in document text — may be computed or misread",
    ),
    "total_taxable_value": FieldResult(value=4380.0, status=FieldStatus.OK),
    "cgst_amount": FieldResult(value=0.0, status=FieldStatus.OK),
    "sgst_amount": FieldResult(value=0.0, status=FieldStatus.OK),
    "igst_amount": FieldResult(value=788.4, status=FieldStatus.OK),
    "total_invoice_value": FieldResult(value=5168.4, status=FieldStatus.OK),
}


@pytest.fixture
def field_configs():
    return load_field_configs()


@pytest.fixture
def sample_results() -> list[DocumentResult]:
    clean = DocumentResult(
        source_filename="SalesBill_AME_2 1.PDF",
        fields={
            "company_name": FieldResult(value="ADITYA MULTIMEDIA AND ENTERTAINMENT", status=FieldStatus.OK),
            "invoice_number": FieldResult(value="AME/2", status=FieldStatus.OK),
            "invoice_date": FieldResult(
                value={"original": "23/04/2026", "normalized": "2026-04-23"},
                status=FieldStatus.OK,
            ),
            "supplier_gstin": FieldResult(value="24ABMFA4190N1Z2", status=FieldStatus.OK),
            "recipient_gstin": FieldResult(status=FieldStatus.NOT_FOUND, reason="field not extracted"),
            "pan": FieldResult(value="ABMFA4190N", status=FieldStatus.OK),
            "description": FieldResult(value="ROYALTY INCOME", status=FieldStatus.OK),
            "total_taxable_value": FieldResult(value=4380.0, status=FieldStatus.OK),
            "cgst_amount": FieldResult(value=0.0, status=FieldStatus.OK),
            "sgst_amount": FieldResult(value=0.0, status=FieldStatus.OK),
            "igst_amount": FieldResult(value=788.4, status=FieldStatus.OK),
            "total_invoice_value": FieldResult(value=5168.4, status=FieldStatus.OK),
        },
        overall_status=DocumentStatus.PARTIAL,
    )

    low_confidence = DocumentResult(
        source_filename="flagged_description.pdf",
        fields=dict(SMOKE_FIELDS),
        overall_status=DocumentStatus.PARTIAL,
    )

    failed_validation = DocumentResult(
        source_filename="bad_gstin.pdf",
        fields={
            **{k: v for k, v in SMOKE_FIELDS.items() if k != "supplier_gstin"},
            "supplier_gstin": FieldResult(
                value="24ABFMA4190N1Z2",
                status=FieldStatus.FAILED_VALIDATION,
                reason="GSTIN checksum failed",
            ),
        },
        overall_status=DocumentStatus.PARTIAL,
    )

    fully_failed = DocumentResult(
        source_filename="corrupt.pdf",
        fields={
            key: FieldResult(status=FieldStatus.NOT_FOUND, reason="processing failed")
            for key in SMOKE_FIELDS
        },
        overall_status=DocumentStatus.FAILED,
        error_message="Failed to load PDF corrupt.pdf: not a real pdf",
    )

    return [clean, low_confidence, failed_validation, fully_failed]


def _expected_headers(field_configs) -> list[str]:
    return (
        ["Source File"]
        + [field.display_label for field in field_configs]
        + ["Overall Status"]
    )


def test_excel_headers_and_clean_row_values(tmp_path: Path, field_configs, sample_results) -> None:
    out = write_excel(sample_results, field_configs, tmp_path / "out.xlsx")
    workbook = load_workbook(out)
    worksheet = workbook.active

    headers = [cell.value for cell in worksheet[1]]
    assert headers == _expected_headers(field_configs)

    clean_row = [cell.value for cell in worksheet[2]]
    assert clean_row[0] == "SalesBill_AME_2 1.PDF"
    assert clean_row[headers.index("Invoice Date")] == "2026-04-23"
    assert clean_row[headers.index("Total Taxable Value")] == 4380.0
    assert clean_row[headers.index("IGST Amount")] == 788.4
    assert clean_row[-1] == "PARTIAL"


def test_excel_cell_fills_for_flagged_and_failed_rows(
    tmp_path: Path,
    field_configs,
    sample_results,
) -> None:
    out = write_excel(sample_results, field_configs, tmp_path / "out.xlsx")
    workbook = load_workbook(out)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]

    description_col = headers.index("Description") + 1
    low_cell = worksheet.cell(row=3, column=description_col)
    assert low_cell.fill.start_color.rgb in {"00FFEB9C", "FFEB9C"}
    assert low_cell.comment is not None
    assert "not found in document text" in low_cell.comment.text

    gstin_col = headers.index("Supplier GSTIN") + 1
    failed_cell = worksheet.cell(row=4, column=gstin_col)
    assert failed_cell.fill.start_color.rgb in {"00FFC7CE", "FFC7CE"}
    assert failed_cell.comment is not None
    assert "checksum failed" in failed_cell.comment.text

    taxable_col = headers.index("Total Taxable Value") + 1
    failed_doc_cell = worksheet.cell(row=5, column=taxable_col)
    assert failed_doc_cell.value is None
    assert failed_doc_cell.fill.start_color.rgb in {"00D9D9D9", "D9D9D9"}

    overall_col = len(headers)
    failed_overall = worksheet.cell(row=5, column=overall_col)
    assert str(failed_overall.value).startswith("FAILED:")
    assert failed_overall.comment is not None


def test_csv_columns_and_flags_column(tmp_path: Path, field_configs, sample_results) -> None:
    out = write_csv(sample_results, field_configs, tmp_path / "out.csv")
    frame = pd.read_csv(out, keep_default_na=False)

    expected_columns = _expected_headers(field_configs) + ["Flags"]
    assert list(frame.columns) == expected_columns

    clean_row = frame.iloc[0]
    assert clean_row["Source File"] == "SalesBill_AME_2 1.PDF"
    assert clean_row["Invoice Date"] == "2026-04-23"
    assert clean_row["Overall Status"] == "PARTIAL"
    assert "recipient_gstin: NOT_FOUND" in clean_row["Flags"]

    low_row = frame.iloc[1]
    assert "description: LOW_CONFIDENCE" in low_row["Flags"]
    assert "not found in document text" in low_row["Flags"]

    bad_gstin_row = frame.iloc[2]
    assert bad_gstin_row["Supplier GSTIN"] == "24ABFMA4190N1Z2"
    assert "supplier_gstin: FAILED_VALIDATION" in bad_gstin_row["Flags"]

    failed_row = frame.iloc[3]
    assert failed_row["Company Name"] == ""
    assert "document: FAILED" in failed_row["Flags"]
    assert "not a real pdf" in failed_row["Flags"]
